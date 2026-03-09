import os
import io
import re
import json
import unicodedata
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG INICIAL
# =========================================================

st.set_page_config(
    page_title="Buscar Cadastro",
    page_icon="🔎",
    layout="wide"
)

BASE_DIR = Path(__file__).resolve().parent
USERS_FILE = BASE_DIR / "users.json"

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

try:
    import gdown
except Exception:
    gdown = None


def get_secret_or_env(name, default=""):
    try:
        if name in st.secrets:
            return st.secrets[name]
    except Exception:
        pass
    return os.getenv(name, default)


DEFAULT_XLSX_PATH = str(get_secret_or_env("XLSX_PATH", "BUSCAR_CAD.xlsx")).strip()
DEFAULT_DRIVE_FILE_ID = str(get_secret_or_env("DRIVE_FILE_ID", "")).strip()

# =========================================================
# LOGIN / USUÁRIOS
# =========================================================

def normalizar_usuario(nome):
    return str(nome).strip()


def ler_admin_por_chaves_diretas():
    """
    Formato alternativo de secrets:
    ADMIN_USER = "admin"
    ADMIN_PASSWORD = "1234"
    """
    try:
        admin_user = str(get_secret_or_env("ADMIN_USER", "admin")).strip()
        admin_password = str(get_secret_or_env("ADMIN_PASSWORD", "")).strip()
        if admin_password:
            return {
                normalizar_usuario(admin_user or "admin"): {
                    "senha": admin_password,
                    "perfil": "total"
                }
            }
    except Exception:
        pass
    return None


def ler_admin_por_bloco_usuarios():
    """
    Formato esperado:
    [USUARIOS]
    admin = { senha = "1234", perfil = "total" }
    """
    try:
        if "USUARIOS" in st.secrets:
            raw = st.secrets["USUARIOS"]

            # Caso seja dict-like
            if hasattr(raw, "items"):
                raw = dict(raw)

            if isinstance(raw, dict) and "admin" in raw:
                valor = raw["admin"]

                if isinstance(valor, dict):
                    senha = str(valor.get("senha", "")).strip()
                    if senha:
                        return {
                            "admin": {
                                "senha": senha,
                                "perfil": "total"
                            }
                        }

                # Caso venha como string direta
                senha = str(valor).strip()
                if senha and senha != "{}":
                    return {
                        "admin": {
                            "senha": senha,
                            "perfil": "total"
                        }
                    }
    except Exception:
        pass

    return None


def ler_admin_por_json():
    """
    Formato alternativo:
    USUARIOS_JSON = '{"admin":{"senha":"1234","perfil":"total"}}'
    """
    try:
        raw_json = None

        if "USUARIOS_JSON" in st.secrets:
            raw_json = st.secrets["USUARIOS_JSON"]
        elif os.getenv("USUARIOS_JSON", "").strip():
            raw_json = os.getenv("USUARIOS_JSON", "").strip()

        if raw_json:
            data = json.loads(raw_json)
            if isinstance(data, dict) and "admin" in data:
                valor = data["admin"]
                if isinstance(valor, dict):
                    senha = str(valor.get("senha", "")).strip()
                    if senha:
                        return {
                            "admin": {
                                "senha": senha,
                                "perfil": "total"
                            }
                        }
                else:
                    senha = str(valor).strip()
                    if senha:
                        return {
                            "admin": {
                                "senha": senha,
                                "perfil": "total"
                            }
                        }
    except Exception:
        pass

    return None


def carregar_admin_de_secrets():
    """
    Ordem:
    1. ADMIN_USER + ADMIN_PASSWORD
    2. [USUARIOS].admin
    3. USUARIOS_JSON
    4. fallback
    """
    por_chaves = ler_admin_por_chaves_diretas()
    if por_chaves:
        return por_chaves

    por_bloco = ler_admin_por_bloco_usuarios()
    if por_bloco:
        return por_bloco

    por_json = ler_admin_por_json()
    if por_json:
        return por_json

    return {
        "admin": {
            "senha": "1234",
            "perfil": "total"
        }
    }


def carregar_usuarios():
    """
    Regra:
    - admin sempre vem de secrets/env/fallback
    - users.json guarda só usuários criados pelo app
    - users.json não sobrescreve admin
    """
    usuarios = carregar_admin_de_secrets()

    if USERS_FILE.exists():
        try:
            extra = json.loads(USERS_FILE.read_text(encoding="utf-8"))
            if isinstance(extra, dict):
                for user, valor in extra.items():
                    user = normalizar_usuario(user)
                    if not user or user.lower() == "admin":
                        continue

                    if isinstance(valor, dict):
                        senha = str(valor.get("senha", "123")).strip()
                        perfil = str(valor.get("perfil", "consulta")).strip().lower()
                    else:
                        senha = str(valor).strip()
                        perfil = "consulta"

                    usuarios[user] = {
                        "senha": senha,
                        "perfil": "total" if perfil == "total" else "consulta"
                    }
        except Exception:
            pass

    # garantia final
    if "admin" not in usuarios:
        usuarios["admin"] = {"senha": "1234", "perfil": "total"}

    usuarios["admin"]["perfil"] = "total"
    usuarios["admin"]["senha"] = str(usuarios["admin"].get("senha", "1234")).strip()

    return usuarios


def salvar_usuarios(usuarios_dict):
    """
    Salva apenas usuários não-admin.
    """
    dados_para_salvar = {}
    for user, dados in usuarios_dict.items():
        user = normalizar_usuario(user)
        if not user or user.lower() == "admin":
            continue
        dados_para_salvar[user] = {
            "senha": str(dados.get("senha", "123")).strip(),
            "perfil": "total" if str(dados.get("perfil", "consulta")).lower() == "total" else "consulta"
        }

    USERS_FILE.write_text(
        json.dumps(dados_para_salvar, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


USUARIOS = carregar_usuarios()

# =========================================================
# CSS
# =========================================================

st.markdown("""
<style>
.block-container {
    padding-top: 1.2rem;
    padding-bottom: 1rem;
}
.login-wrap {
    max-width: 460px;
    margin: 5vh auto 0 auto;
}
.login-box {
    background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 18px;
    padding: 28px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.25);
}
.login-title {
    font-size: 28px;
    font-weight: 700;
    margin-bottom: 6px;
    text-align: center;
}
.login-subtitle {
    color: #cbd5e1;
    margin-bottom: 20px;
    text-align: center;
}
.small-note {
    color: #94a3b8;
    font-size: 0.9rem;
    text-align: center;
    margin-top: 12px;
}
div[data-testid="stTextInput"] input {
    border-radius: 10px;
}
div[data-testid="stButton"] button {
    border-radius: 10px;
    font-weight: 600;
}
.result-box {
    padding: 12px;
    border-radius: 12px;
    background: rgba(59,130,246,0.08);
    border: 1px solid rgba(59,130,246,0.15);
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# COLUNAS PADRÃO DO RESULTADO / EXPORT
# =========================================================

COLUNAS_PADRAO = [
    "EMPRESA",
    "REGIONAL",
    "BASE",
    "NOME",
    "NOTA AM",
    "ID SAP",
    "DEPOSITO",
    "PN",
    "DESCRIÇÃO",
    "ATENDIDO POR",
    "DATA DA BAIXA",
]

MAPA_CAMPOS = {
    "EMPRESA": ["EMPRESA"],
    "REGIONAL": ["REGIONAL"],
    "BASE": ["BASE"],
    "NOME": ["NOME", "ELETRICISTA", "NOME DO ELETRICISTA", "NOME COMPLETO", "PRESTADOR", "COLABORADOR"],
    "NOTA AM": ["NOTA AM", "NOTA", "NF", "NUMERO DA NOTA", "NUMERO NOTA", "N DA NOTA", "N_NOTA"],
    "ID SAP": ["ID SAP", "IDSAP", "SAP", "ID_SAP"],
    "DEPOSITO": ["DEPOSITO", "DEPÓSITO"],
    "PN": ["PN"],
    "DESCRIÇÃO": ["DESCRICAO", "DESCRIÇÃO", "DESC", "DESCRICAO MATERIAL"],
    "ATENDIDO POR": ["ATENDIDO POR", "ATENDIDO_POR"],
    "DATA DA BAIXA": ["DATA DA BAIXA", "DT BAIXA", "DATA BAIXA", "BAIXA", "DATA"],
}

# =========================================================
# UTILITÁRIOS
# =========================================================

def normalizar_texto(valor):
    if pd.isna(valor):
        return ""
    valor = str(valor).strip()
    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(c for c in valor if not unicodedata.combining(c))
    valor = re.sub(r"\s+", " ", valor)
    return valor.upper().strip()


def slug_coluna(valor):
    txt = normalizar_texto(valor)
    txt = txt.replace("Ç", "C")
    txt = re.sub(r"[^A-Z0-9]+", "_", txt)
    txt = re.sub(r"_+", "_", txt).strip("_")
    return txt


def safe_str_contains(series, termo):
    termo_norm = normalizar_texto(termo)
    serie_norm = series.fillna("").astype(str).map(normalizar_texto)
    return serie_norm.str.contains(re.escape(termo_norm), na=False)


def resolver_caminho_xlsx(path_xlsx):
    p = Path(str(path_xlsx).strip())
    if not p.is_absolute():
        p = BASE_DIR / p
    return p.resolve()


def detectar_cabecalho(path_xlsx, sheet_name, max_linhas=12):
    bruto = pd.read_excel(path_xlsx, sheet_name=sheet_name, header=None, nrows=max_linhas)
    melhor_idx = 0
    melhor_score = -1

    chaves_nota = ["NOTA", "NF", "NUMERO_NOTA", "N_NOTA", "NOTA_AM"]
    chaves_data = ["DATA", "DT", "DATA_BAIXA"]
    chaves_nome = ["ELETRICISTA", "NOME", "NOME_ELETRICISTA", "PRESTADOR", "COLABORADOR"]

    for idx in range(len(bruto)):
        linha = [slug_coluna(v) for v in bruto.iloc[idx].tolist()]
        score = 0
        for c in linha:
            if any(k in c for k in chaves_nota):
                score += 3
            if any(k in c for k in chaves_data):
                score += 2
            if any(k in c for k in chaves_nome):
                score += 4
        if score > melhor_score:
            melhor_score = score
            melhor_idx = idx

    return melhor_idx


def localizar_coluna(df, candidatos):
    mapa = {slug_coluna(c): c for c in df.columns}

    for cand in candidatos:
        cand_slug = slug_coluna(cand)
        if cand_slug in mapa:
            return mapa[cand_slug]

    for slug, original in mapa.items():
        for cand in candidatos:
            cand_slug = slug_coluna(cand)
            if cand_slug in slug:
                return original

    return None


def limpar_dataframe(df):
    df = df.dropna(axis=1, how="all").copy()

    novas_cols = []
    for i, col in enumerate(df.columns):
        col_str = str(col).strip()
        if not col_str or col_str.upper().startswith("UNNAMED"):
            col_str = f"COLUNA_{i+1}"
        novas_cols.append(col_str)
    df.columns = novas_cols

    df = df.dropna(axis=0, how="all").copy()
    return df


def preparar_dataframe(path_xlsx, sheet_name):
    caminho_real = resolver_caminho_xlsx(path_xlsx)
    header_idx = detectar_cabecalho(caminho_real, sheet_name)
    df = pd.read_excel(caminho_real, sheet_name=sheet_name, header=header_idx)
    df = limpar_dataframe(df)

    if len(df) > 0:
        primeira = [normalizar_texto(x) for x in df.iloc[0].tolist()]
        cab = [normalizar_texto(x) for x in df.columns.tolist()]
        if primeira == cab:
            df = df.iloc[1:].copy()

    return df, header_idx


def listar_abas(path_xlsx):
    caminho_real = resolver_caminho_xlsx(path_xlsx)
    xls = pd.ExcelFile(caminho_real)
    return xls.sheet_names


def baixar_do_drive(file_id, destino):
    if not file_id:
        return False, "DRIVE_FILE_ID não informado."
    if gdown is None:
        return False, "Biblioteca gdown não instalada. Adicione gdown no requirements.txt."

    destino_real = resolver_caminho_xlsx(destino)
    destino_real.parent.mkdir(parents=True, exist_ok=True)

    url = f"https://drive.google.com/uc?id={file_id}"
    try:
        gdown.download(url, str(destino_real), quiet=False)
        return True, f"Arquivo baixado com sucesso para: {destino_real}"
    except Exception as e:
        return False, f"Erro ao baixar do Drive: {e}"


def validar_xlsx(path_xlsx):
    p = resolver_caminho_xlsx(path_xlsx)

    if not p.exists():
        return False, f"Arquivo não encontrado: {p}"

    if p.suffix.lower() not in [".xlsx", ".xlsm", ".xls"]:
        return False, f"Arquivo inválido: {p}"

    return True, str(p)


def montar_resultado_padrao(df):
    resultado = pd.DataFrame()

    for coluna_final in COLUNAS_PADRAO:
        coluna_origem = localizar_coluna(df, MAPA_CAMPOS[coluna_final])
        if coluna_origem:
            resultado[coluna_final] = df[coluna_origem]
        else:
            resultado[coluna_final] = ""

    if "DATA DA BAIXA" in resultado.columns:
        serie_data = pd.to_datetime(resultado["DATA DA BAIXA"], errors="coerce")
        resultado["DATA DA BAIXA"] = serie_data.dt.strftime("%d/%m/%Y").fillna("")

    for c in resultado.columns:
        resultado[c] = resultado[c].fillna("").astype(str)

    return resultado


def df_para_excel_bytes_formatado(df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="resultado")
        ws = writer.book["resultado"]

        fill_azul = PatternFill("solid", fgColor="9DC3E6")
        fill_verde = PatternFill("solid", fgColor="C6E0B4")
        fonte = Font(bold=True, color="000000")
        borda = Border(
            left=Side(style="thin", color="808080"),
            right=Side(style="thin", color="808080"),
            top=Side(style="thin", color="808080"),
            bottom=Side(style="thin", color="808080")
        )
        alinhamento = Alignment(horizontal="center", vertical="center")

        for idx, cell in enumerate(ws[1], start=1):
            cell.font = fonte
            cell.border = borda
            cell.alignment = alinhamento
            cell.fill = fill_verde if idx >= 8 else fill_azul

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = borda

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            for value in df.iloc[:, col_idx - 1].astype(str).head(1000):
                if len(value) > max_len:
                    max_len = len(value)
            largura = min(max(max_len + 2, 12), 35)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

    output.seek(0)
    return output.getvalue()


def gerar_texto_notas(df):
    if "NOTA AM" not in df.columns or df.empty:
        return ""
    notas = df["NOTA AM"].fillna("").astype(str).str.strip()
    notas = [n for n in notas if n]
    return "\n".join(notas)


def desempacotar_carregamento(retorno):
    if isinstance(retorno, tuple) and len(retorno) >= 2:
        return retorno[0], retorno[1]
    raise ValueError("Retorno inesperado ao carregar a planilha.")


def usuario_e_admin():
    return str(st.session_state.get("usuario_logado", "")).strip().lower() == "admin"


def perfil_usuario_logado():
    usuario = str(st.session_state.get("usuario_logado", "")).strip()
    dados = USUARIOS.get(usuario, {})
    perfil = str(dados.get("perfil", "consulta")).strip().lower()
    return "total" if perfil == "total" else "consulta"


def usuario_pode_configurar():
    return perfil_usuario_logado() == "total"


def recarregar_usuarios():
    global USUARIOS
    USUARIOS = carregar_usuarios()


def tentar_baixar_automaticamente_se_faltar():
    ok, _ = validar_xlsx(st.session_state.xlsx_path)
    if ok:
        return

    if st.session_state.usar_drive and st.session_state.drive_file_id:
        baixar_do_drive(st.session_state.drive_file_id, st.session_state.xlsx_path)

# =========================================================
# ESTADO
# =========================================================

if "logado" not in st.session_state:
    st.session_state.logado = False

if "usuario_logado" not in st.session_state:
    st.session_state.usuario_logado = ""

if "xlsx_path" not in st.session_state:
    st.session_state.xlsx_path = DEFAULT_XLSX_PATH

if "drive_file_id" not in st.session_state:
    st.session_state.drive_file_id = DEFAULT_DRIVE_FILE_ID

if "usar_drive" not in st.session_state:
    st.session_state.usar_drive = True if DEFAULT_DRIVE_FILE_ID else False

# =========================================================
# CACHE
# =========================================================

@st.cache_data(show_spinner=False)
def carregar_abas_cache(path_xlsx):
    return listar_abas(path_xlsx)


@st.cache_data(show_spinner=False)
def carregar_df_cache(path_xlsx, sheet_name):
    return preparar_dataframe(path_xlsx, sheet_name)

# =========================================================
# LOGIN
# =========================================================

def tela_login():
    st.markdown('<div class="login-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="login-box">', unsafe_allow_html=True)
    st.markdown('<div class="login-title">🔐 Buscar Cadastro</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-subtitle">Entre com seu usuário e senha para acessar o sistema.</div>', unsafe_allow_html=True)

    with st.form("form_login", clear_on_submit=False):
        usuario = st.text_input("Usuário", key="login_usuario")
        senha = st.text_input("Senha", type="password", key="login_senha")
        entrar = st.form_submit_button("Entrar", use_container_width=True)

    if entrar:
        recarregar_usuarios()

        usuario = normalizar_usuario(usuario)
        senha_digitada = str(senha).strip()

        usuario_ok = usuario in USUARIOS
        senha_ok = usuario_ok and str(USUARIOS[usuario]["senha"]).strip() == senha_digitada

        if usuario_ok and senha_ok:
            st.session_state.logado = True
            st.session_state.usuario_logado = usuario
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")
            with st.expander("Diagnóstico do login", expanded=False):
                st.write("Usuário informado:", usuario)
                st.write("Usuário existe:", "sim" if usuario_ok else "não")
                if usuario_ok:
                    st.write("Perfil lido:", USUARIOS[usuario].get("perfil", "consulta"))
                    st.write("Fonte do admin:", "secrets/env/fallback" if usuario.lower() == "admin" else "users.json")
                    st.write("Senha vazia no cadastro:", "sim" if not str(USUARIOS[usuario].get("senha", "")).strip() else "não")

    st.markdown('<div class="small-note">O usuário admin sempre obedece ao Secrets.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# =========================================================
# GESTÃO DE USUÁRIOS
# =========================================================

def painel_usuarios_admin():
    global USUARIOS

    if not usuario_e_admin():
        return

    with st.sidebar:
        with st.expander("👥 Usuários", expanded=False):
            st.caption("Somente o admin pode criar e gerenciar usuários.")

            with st.form("form_novo_usuario", clear_on_submit=True):
                novo_usuario = st.text_input("Novo usuário")
                nova_senha = st.text_input("Senha do novo usuário", type="password")
                novo_perfil = st.selectbox("Perfil", ["consulta", "total"])
                criar_usuario = st.form_submit_button("Criar usuário", use_container_width=True)

            if criar_usuario:
                u = normalizar_usuario(novo_usuario)
                s = str(nova_senha).strip()
                p = str(novo_perfil).strip().lower()

                if not u:
                    st.error("Informe o nome do usuário.")
                elif u.lower() == "admin":
                    st.error("O usuário admin é controlado pelo Secrets e não pode ser recriado aqui.")
                elif not re.match(r"^[A-Za-z0-9_.-]+$", u):
                    st.error("Usuário inválido. Use apenas letras, números, _, . ou -")
                elif not s:
                    st.error("Informe a senha.")
                elif u in USUARIOS:
                    st.error("Este usuário já existe.")
                else:
                    USUARIOS[u] = {"senha": s, "perfil": "total" if p == "total" else "consulta"}
                    try:
                        salvar_usuarios(USUARIOS)
                        recarregar_usuarios()
                        st.success(f"Usuário '{u}' criado com sucesso.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Não foi possível salvar o usuário: {e}")

            usuarios_exibicao = []
            for user, dados in USUARIOS.items():
                usuarios_exibicao.append({
                    "USUÁRIO": user,
                    "PERFIL": dados.get("perfil", "consulta")
                })

            if usuarios_exibicao:
                st.dataframe(pd.DataFrame(usuarios_exibicao), use_container_width=True, height=220)

            st.caption("Alteração rápida de perfil/senha")

            usuarios_editaveis = list(USUARIOS.keys())
            if usuarios_editaveis:
                with st.form("form_editar_usuario", clear_on_submit=True):
                    usuario_alvo = st.selectbox("Selecionar usuário", usuarios_editaveis)
                    nova_senha_edit = st.text_input("Nova senha (deixe em branco para manter)", type="password")
                    novo_perfil_edit = st.selectbox("Novo perfil", ["consulta", "total"])
                    salvar_edicao = st.form_submit_button("Salvar alterações", use_container_width=True)

                if salvar_edicao:
                    try:
                        if usuario_alvo not in USUARIOS:
                            st.error("Usuário não encontrado.")
                        elif usuario_alvo.lower() == "admin":
                            st.error("O admin deve ser alterado pelo Secrets do Streamlit Cloud.")
                        else:
                            if nova_senha_edit.strip():
                                USUARIOS[usuario_alvo]["senha"] = nova_senha_edit.strip()
                            USUARIOS[usuario_alvo]["perfil"] = "total" if novo_perfil_edit == "total" else "consulta"

                            salvar_usuarios(USUARIOS)
                            recarregar_usuarios()
                            st.success("Usuário atualizado com sucesso.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao atualizar usuário: {e}")

                with st.form("form_excluir_usuario", clear_on_submit=True):
                    usuarios_para_excluir = [u for u in usuarios_editaveis if u.lower() != "admin"]
                    if usuarios_para_excluir:
                        usuario_excluir = st.selectbox("Excluir usuário", usuarios_para_excluir)
                        excluir = st.form_submit_button("Excluir usuário", use_container_width=True)
                        if excluir:
                            try:
                                USUARIOS.pop(usuario_excluir, None)
                                salvar_usuarios(USUARIOS)
                                recarregar_usuarios()
                                st.success(f"Usuário '{usuario_excluir}' removido.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro ao excluir usuário: {e}")
                    else:
                        st.info("Não há usuários removíveis no momento.")

# =========================================================
# SIDEBAR / CONFIGURAÇÕES
# =========================================================

def painel_configuracoes():
    pode_configurar = usuario_pode_configurar()

    with st.sidebar:
        st.markdown("## ⚙️ Configurações")
        st.write(f"Usuário logado: **{st.session_state.usuario_logado}**")
        st.write(f"Perfil: **{perfil_usuario_logado()}**")

        if pode_configurar:
            st.success("Acesso total")
        else:
            st.info("Acesso consulta")
            st.caption("Somente usuários com perfil total podem alterar as configurações.")

        with st.expander("Arquivo XLSX", expanded=False):
            if pode_configurar:
                usar_drive = st.checkbox(
                    "Baixar do Google Drive",
                    value=st.session_state.usar_drive
                )

                xlsx_path = st.text_input(
                    "Caminho do XLSX",
                    value=st.session_state.xlsx_path,
                    help="Ex.: BUSCAR_CAD.xlsx ou subpasta/BUSCAR_CAD.xlsx"
                )

                drive_file_id = st.text_input(
                    "Drive File ID",
                    value=st.session_state.drive_file_id,
                    help="Use se quiser baixar a planilha direto do Google Drive."
                )

                col_a, col_b = st.columns(2)

                with col_a:
                    if st.button("Salvar config", use_container_width=True):
                        st.session_state.usar_drive = usar_drive
                        st.session_state.xlsx_path = xlsx_path.strip() or DEFAULT_XLSX_PATH
                        st.session_state.drive_file_id = drive_file_id.strip()
                        st.success("Configurações salvas.")
                        st.rerun()

                with col_b:
                    if st.button("Restaurar padrão", use_container_width=True):
                        st.session_state.usar_drive = True if DEFAULT_DRIVE_FILE_ID else False
                        st.session_state.xlsx_path = DEFAULT_XLSX_PATH
                        st.session_state.drive_file_id = DEFAULT_DRIVE_FILE_ID
                        st.success("Padrão restaurado.")
                        st.rerun()

                if st.button("Baixar/atualizar do Drive", use_container_width=True):
                    destino = st.session_state.xlsx_path
                    ok, msg = baixar_do_drive(st.session_state.drive_file_id, destino)
                    if ok:
                        carregar_abas_cache.clear()
                        carregar_df_cache.clear()
                        st.success(msg)
                    else:
                        st.error(msg)
            else:
                st.text_input("Caminho do XLSX", value=st.session_state.xlsx_path, disabled=True)
                st.text_input("Drive File ID", value=st.session_state.drive_file_id, disabled=True)
                st.checkbox("Baixar do Google Drive", value=st.session_state.usar_drive, disabled=True)
                st.warning("Configurações bloqueadas para este usuário.")

        try:
            tentar_baixar_automaticamente_se_faltar()
        except Exception:
            pass

        ok, msg = validar_xlsx(st.session_state.xlsx_path)
        if ok:
            st.success(f"Arquivo ativo: {msg}")
        else:
            st.warning("Arquivo ativo não encontrado.")
            st.caption(msg)

        with st.expander("Diagnóstico do caminho", expanded=False):
            st.write("Pasta do app:", str(BASE_DIR))
            st.write("Caminho configurado:", st.session_state.xlsx_path)
            st.write("Caminho resolvido:", str(resolver_caminho_xlsx(st.session_state.xlsx_path)))
            st.write("Usar Drive:", st.session_state.usar_drive)
            st.write("Drive File ID informado:", "sim" if st.session_state.drive_file_id else "não")

        with st.expander("Diagnóstico do admin", expanded=False):
            admin_info = USUARIOS.get("admin", {})
            st.write("Admin carregado:", "sim" if "admin" in USUARIOS else "não")
            st.write("Perfil admin:", admin_info.get("perfil", ""))
            st.write("Senha admin vazia:", "sim" if not str(admin_info.get("senha", "")).strip() else "não")
            st.write("Formato detectado de secrets:", 
                     "ADMIN_USER/ADMIN_PASSWORD"
                     if ler_admin_por_chaves_diretas()
                     else "[USUARIOS].admin"
                     if ler_admin_por_bloco_usuarios()
                     else "USUARIOS_JSON"
                     if ler_admin_por_json()
                     else "fallback")

        if pode_configurar:
            if st.button("Limpar cache", use_container_width=True):
                carregar_abas_cache.clear()
                carregar_df_cache.clear()
                st.success("Cache limpo. Recarregue a busca.")
        else:
            st.button("Limpar cache", use_container_width=True, disabled=True)

        if st.button("Sair", use_container_width=True):
            st.session_state.logado = False
            st.session_state.usuario_logado = ""
            st.rerun()

# =========================================================
# APP PRINCIPAL
# =========================================================

def app():
    painel_configuracoes()
    painel_usuarios_admin()

    st.title("🔎 Buscar Cadastro de Eletricista")

    caminho = st.session_state.xlsx_path
    ok, msg = validar_xlsx(caminho)
    if not ok:
        st.error(f"Não foi possível abrir o XLSX. Motivo: {msg}")
        st.stop()

    caminho = msg

    try:
        abas = carregar_abas_cache(caminho)
    except Exception as e:
        st.error(f"Erro ao listar abas da planilha: {e}")
        st.stop()

    if not abas:
        st.error("Nenhuma aba encontrada no arquivo.")
        st.stop()

    abas_opcoes = abas + ["TODOS"]

    top1, top2 = st.columns([1.1, 2.2])

    with top1:
        aba_escolhida = st.selectbox("Selecione a aba", abas_opcoes)

    with top2:
        st.info("O seletor mostra automaticamente as abas existentes na planilha.")

    if aba_escolhida == "TODOS":
        dfs = []
        for aba in abas:
            try:
                retorno = carregar_df_cache(caminho, aba)
                df_tmp, _ = desempacotar_carregamento(retorno)
                if len(df_tmp) > 0:
                    dfs.append(df_tmp)
            except Exception:
                continue

        if not dfs:
            st.error("Não foi possível carregar nenhuma aba válida.")
            st.stop()

        df_bruto = pd.concat(dfs, ignore_index=True)
    else:
        try:
            retorno = carregar_df_cache(caminho, aba_escolhida)
            df_bruto, _ = desempacotar_carregamento(retorno)
        except Exception as e:
            st.error(f"Erro ao carregar a aba '{aba_escolhida}': {e}")
            st.info("Se você acabou de trocar a versão do app, clique em 'Limpar cache' e tente novamente.")
            st.stop()

    df = montar_resultado_padrao(df_bruto)

    st.divider()

    c1, c2 = st.columns(2)

    with c1:
        busca_nome = st.text_input("Digite o nome")
        busca_exata = st.checkbox("Busca exata do nome", value=False)

    with c2:
        busca_nota = st.text_input("Digite a nota AM")

    usar_data = st.checkbox("Filtrar por período da data da baixa", value=False)
    data_inicial = None
    data_final = None

    if usar_data:
        d1, d2 = st.columns(2)
        with d1:
            data_inicial = st.date_input("Data inicial")
        with d2:
            data_final = st.date_input("Data final")

    st.divider()

    st.subheader("Busca em massa por nomes")
    lista_nomes = st.text_area(
        "Cole vários nomes (1 por linha)",
        height=160,
        help="Exemplo: um nome por linha. O sistema vai buscar todos de uma vez."
    )

    b1, b2 = st.columns(2)
    pesquisar = b1.button("🔎 Pesquisar", use_container_width=True)
    limpar = b2.button("Limpar filtros", use_container_width=True)

    if limpar:
        st.rerun()

    resultado = df.copy()

    if pesquisar:
        if busca_nome:
            serie_nome = resultado["NOME"].fillna("").astype(str)
            if busca_exata:
                nome_ref = normalizar_texto(busca_nome)
                resultado = resultado[serie_nome.map(normalizar_texto) == nome_ref]
            else:
                resultado = resultado[safe_str_contains(serie_nome, busca_nome)]

        if busca_nota:
            serie_nota = resultado["NOTA AM"].fillna("").astype(str).str.strip()
            nota_ref = str(busca_nota).strip()
            resultado = resultado[serie_nota.str.contains(re.escape(nota_ref), na=False)]

        if usar_data and data_inicial and data_final:
            serie_data = pd.to_datetime(resultado["DATA DA BAIXA"], format="%d/%m/%Y", errors="coerce")
            resultado = resultado[
                (serie_data.dt.date >= data_inicial) &
                (serie_data.dt.date <= data_final)
            ]

        if lista_nomes.strip():
            nomes = [linha.strip() for linha in lista_nomes.splitlines() if linha.strip()]
            nomes_normalizados = {normalizar_texto(nome) for nome in nomes}
            serie_nome_norm = resultado["NOME"].fillna("").astype(str).map(normalizar_texto)
            resultado = resultado[serie_nome_norm.isin(nomes_normalizados)]

    st.divider()

    if pesquisar:
        if resultado.empty:
            st.warning("Nenhum registro encontrado.")
        else:
            st.success(f"{len(resultado)} registro(s) encontrado(s).")

            with st.container():
                st.markdown('<div class="result-box">Resultado da pesquisa carregado com sucesso.</div>', unsafe_allow_html=True)

            st.dataframe(resultado[COLUNAS_PADRAO], use_container_width=True, height=420)

            col_btn1, col_btn2 = st.columns(2)

            with col_btn1:
                excel_bytes = df_para_excel_bytes_formatado(resultado[COLUNAS_PADRAO])
                st.download_button(
                    "⬇️ Baixar resultado em Excel",
                    data=excel_bytes,
                    file_name="resultado_busca_cadastro.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col_btn2:
                notas_texto = gerar_texto_notas(resultado[COLUNAS_PADRAO])
                st.code(notas_texto if notas_texto else "", language=None)
                st.caption("Copie o conteúdo acima. Ele traz as notas da coluna NOTA AM, uma por linha.")

    with st.expander("📄 Visualizar base completa", expanded=False):
        st.dataframe(df_bruto, use_container_width=True, height=350)
        st.caption("Colunas encontradas na base original:")
        st.write(list(df_bruto.columns))

# =========================================================
# MAIN
# =========================================================

def main():
    if not st.session_state.logado:
        tela_login()
    else:
        app()

if __name__ == "__main__":
    main()
